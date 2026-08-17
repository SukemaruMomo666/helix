import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

products = [
    {"nama": "Akar kelapa (1kg)", "harga_jual": 50000},
    {"nama": "Kembang goyang (isi 100)", "harga_jual": 50000},
    {"nama": "Kiripik pisang (1kg)", "harga_jual": 50000},
    {"nama": "Opak (isi 100)", "harga_jual": 50000},
    {"nama": "Pungpa (1kg)", "harga_jual": 50000},
    {"nama": "rengginang (isi 100)", "harga_jual": 80000},
    {"nama": "Sistik Asli (1kg)", "harga_jual": 50000},
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
fill_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

font_white_bold_italic = Font(color="FFFFFF", bold=True, italic=True)
font_white_bold = Font(color="FFFFFF", bold=True)
font_black_bold_italic = Font(color="000000", bold=True, italic=True)
font_black_italic = Font(color="000000", italic=True)

align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_currency(val):
    return f"Rp{val:,.0f}"

for prod in products:
    # Limit sheet title to 31 chars
    sheet_title = prod['nama'][:31].replace('/', '_').replace(':', '')
    ws = wb.create_sheet(title=sheet_title)
    
    hj = prod['harga_jual']
    
    # We will simulate HPP so that Cost of Sales + Margin = Harga Jual
    # Let's target Margin = 50% of Cost of Sales, so Harga Jual = 1.5 * Cost of Sales
    # Cost of Sales = Harga Jual / 1.5
    cost_of_sales = int(hj / 1.5)
    margin_val = hj - cost_of_sales
    
    # Platform = 10% of Harga Jual
    biaya_platform = int(hj * 0.10)
    
    # Packaging = 2000
    packaging = 2000
    label = 500
    
    # HPP = Cost of Sales - platform - packaging - label
    hpp = cost_of_sales - biaya_platform - packaging - label
    
    # Tenaga Kerja = 5000
    tenaga_kerja = 5000
    
    # Cost of Material = HPP - tenaga_kerja
    cost_material = hpp - tenaga_kerja
    
    # Materials breakdown (approximate percentages)
    bahan_utama = int(cost_material * 0.6)
    minyak = int(cost_material * 0.3)
    bumbu = cost_material - bahan_utama - minyak
    
    rows = [
        ["HPP", "", "", ""],
        [prod['nama'], "", "", ""],
        ["Nama", "Jumlah", "Harga Per UOM", "Harga per Product"],
        ["Bahan Baku Utama", 1, bahan_utama, bahan_utama],
        ["Minyak Goreng", 1, minyak, minyak],
        ["Bumbu / Pelengkap", 1, bumbu, bumbu],
        ["Cost of Material (Bahan Baku)", "", "", cost_material],
        ["Biaya Tenaga Kerja", 1, tenaga_kerja, tenaga_kerja],
        ["Harga Pokok Penjualan (HPP)", "", "", hpp],
        ["Plastik Kemasan", 1, packaging, packaging],
        ["Label Stiker", 1, label, label],
        ["Biaya Platform (e-commerce)", 1, f"10%", biaya_platform],
        ["Cost of Sales", "", "", cost_of_sales],
        ["Margin", "", "50%", margin_val],
        ["Harga Jual", "", "", hj]
    ]
    
    for r_idx, row_data in enumerate(rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            
            if r_idx == 1:
                cell.fill = fill_black
                cell.font = font_white_bold_italic
                cell.alignment = align_center
            elif r_idx == 2:
                cell.fill = fill_green
                cell.font = font_black_bold_italic
                cell.alignment = align_center
            elif r_idx == 3:
                cell.fill = fill_gray
                cell.font = font_white_bold
                cell.alignment = align_center
            elif r_idx in [7, 9, 13]: # Subtotals
                cell.fill = fill_black
                cell.font = font_white_bold_italic
                if c_idx == 1:
                    cell.alignment = align_center
                elif c_idx == 4:
                    cell.alignment = align_right
                    cell.value = format_currency(val)
            elif r_idx == 14: # Margin
                cell.font = font_black_italic
                if c_idx == 1:
                    cell.alignment = align_center
                elif c_idx == 4:
                    cell.value = format_currency(val)
            elif r_idx == 15: # Harga Jual
                cell.fill = fill_green
                cell.font = font_black_bold_italic
                if c_idx == 1:
                    cell.alignment = align_center
                elif c_idx == 4:
                    cell.value = format_currency(val)
            else: # Normal rows
                if c_idx == 1:
                    cell.alignment = align_center
                elif c_idx in [3, 4] and isinstance(val, (int, float)):
                    cell.value = format_currency(val)
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center

    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('A7:C7')
    ws.merge_cells('A9:C9')
    ws.merge_cells('A13:C13')
    ws.merge_cells('A15:C15')
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

wb.save("Template_HPP_Semua_Produk.xlsx")
print("Template_HPP_Semua_Produk.xlsx created.")
